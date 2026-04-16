"""
スキルシート変換ツール
- Excel/PDFのスキルシートをアップロード
- Gemini APIで情報抽出
- 自社レイアウト（SAP職務経歴書）形式で表示・Excel出力
"""

import streamlit as st
import google.generativeai as genai
import openpyxl
import pdfplumber
import fitz  # PyMuPDF
import json
import os
import re
from io import BytesIO
from datetime import datetime
# PDF出力は不使用

# ─────────────────────────────────────────
# 設定
# ─────────────────────────────────────────
# APIキー：ローカル(.env) / Streamlit Cloud(secrets)両対応
try:
    GEMINI_API_KEY = st.secrets["GEMINI_API_KEY"]
except Exception:
    GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")


# ─────────────────────────────────────────
# ページ設定
# ─────────────────────────────────────────
st.set_page_config(
    page_title="スキルシート変換ツール",
    page_icon="📄",
    layout="wide"
)

st.markdown("""
<style>
    .main-title {
        font-size: 2rem;
        font-weight: bold;
        color: #1a3a6b;
        margin-bottom: 0.5rem;
    }
    .section-header {
        background: #1a3a6b;
        color: white;
        padding: 6px 12px;
        border-radius: 4px;
        font-weight: bold;
        margin: 1rem 0 0.5rem 0;
    }
    .project-card {
        background: #f8f9fa;
        border-left: 4px solid #1a3a6b;
        padding: 10px 15px;
        margin: 8px 0;
        border-radius: 0 4px 4px 0;
    }
    .badge {
        background: #1a3a6b;
        color: white;
        padding: 2px 8px;
        border-radius: 10px;
        font-size: 0.8rem;
        margin-right: 4px;
    }
    .skill-tag {
        background: #e8f0fe;
        color: #1a3a6b;
        padding: 2px 8px;
        border-radius: 10px;
        font-size: 0.85rem;
        margin: 2px;
        display: inline-block;
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────
# ユーティリティ関数
# ─────────────────────────────────────────

def extract_text_from_excel(file_bytes: bytes, ext: str = "xlsx") -> str:
    """ExcelファイルからテキストをGrid形式で抽出"""
    result = []
    if ext == "xls":
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            result.append(f"=== シート: {sheet_name} ===")
            for row_idx in range(ws.nrows):
                row_texts = []
                for col_idx in range(ws.ncols):
                    cell = ws.cell(row_idx, col_idx)
                    val = cell.value
                    if val is not None and val != "":
                        col_letter = chr(ord('A') + col_idx) if col_idx < 26 else f"C{col_idx}"
                        row_texts.append(f"[{col_letter}{row_idx+1}]{str(val).strip()}")
                if row_texts:
                    result.append("  |  ".join(row_texts))
    else:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result.append(f"=== シート: {sheet_name} ===")
            for row in ws.iter_rows():
                row_texts = []
                for cell in row:
                    val = cell.value
                    if val is not None:
                        if isinstance(val, datetime):
                            val = val.strftime("%Y/%m")
                        row_texts.append(f"[{cell.coordinate}]{str(val).strip()}")
                if row_texts:
                    result.append("  |  ".join(row_texts))
    return "\n".join(result)


def extract_text_from_pdf(file_bytes: bytes) -> str:
    """PDFからテキストを抽出"""
    text_parts = []
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                text_parts.append(f"=== ページ {i+1} ===\n{text}")
    return "\n".join(text_parts)


def extract_skills_with_gemini(file_text: str, api_key: str) -> dict:
    """Gemini APIでスキルシート情報をJSON抽出"""
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel("gemini-2.5-flash")

    prompt = f"""
以下はスキルシート（職務経歴書）のテキストデータです。
このデータから必要な情報を抽出して、必ず以下のJSON形式のみで返してください。
説明文や```json```の囲みは不要です。JSONオブジェクトだけを返してください。

抽出するJSON形式:
{{
  "基本情報": {{
    "氏名": "",
    "フリガナ": "",
    "性別": "",
    "生年月日": "",
    "年齢": "",
    "未既婚": "",
    "国籍": "",
    "日本滞在年数": "",
    "住所": "",
    "最寄駅路線": "",
    "最寄駅名": ""
  }},
  "SAP情報": {{
    "モジュール": "",
    "ポジション": "",
    "SAP経験年数": ""
  }},
  "言語能力": {{
    "日本語": {{
      "日常会話": "",
      "業務会話": "",
      "読み": "",
      "書き": "",
      "仕様書読解": "",
      "仕様書作成": ""
    }},
    "英語": {{
      "日常会話": "",
      "業務会話": "",
      "読み": "",
      "書き": "",
      "仕様書読解": "",
      "仕様書作成": ""
    }}
  }},
  "取得資格": "",
  "得意分野": "",
  "自己PR": "",
  "職務経歴": [
    {{
      "No": 1,
      "開始年月": "",
      "終了年月": "",
      "業種": "",
      "プロジェクト概要": "",
      "担当業務": "",
      "OS_DB": "",
      "作業環境": "",
      "開発言語": "",
      "役割": "",
      "フェーズ": {{
        "分析調査": false,
        "提案管理レビュー": false,
        "要件定義": false,
        "基本設計": false,
        "詳細設計": false,
        "製造": false,
        "単体試験": false,
        "結合試験": false,
        "総合試験": false,
        "運用保守": false
      }}
    }}
  ]
}}

スキルシートデータ:
{file_text[:8000]}
"""

    response = model.generate_content(prompt)
    raw = response.text.strip()
    # JSONの抽出（```json...``` ブロックがあれば除去）
    raw = re.sub(r"^```json\s*", "", raw)
    raw = re.sub(r"^```\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)



# ─────────────────────────────────────────
# ユーティリティ：None安全文字列変換
# ─────────────────────────────────────────
def safe_str(val, default="-"):
    """NoneやNaN等を安全に文字列変換"""
    if val is None:
        return default
    s = str(val).strip()
    return s if s else default


# ─────────────────────────────────────────
# Excel出力
# ─────────────────────────────────────────
def generate_excel(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.page import PageMargins
    from io import BytesIO as _BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "職務経歴書"

    # グリッド線を非表示（記入エリア外をすっきり見せる）
    ws.sheet_view.showGridLines = False

    # ── スタイル定義 ──────────────────────────────────────
    NAVY      = "1A3A6B"
    WHITE     = "FFFFFF"
    STRIPE_C  = "EEF2FF"

    HDR_FILL  = PatternFill("solid", fgColor=NAVY)
    STR_FILL  = PatternFill("solid", fgColor=STRIPE_C)
    WHT_FILL  = PatternFill("solid", fgColor=WHITE)

    HDR_FONT  = Font(name="Meiryo UI", color=WHITE,  bold=True, size=9)
    VAL_FONT  = Font(name="Meiryo UI", color="000000", size=9)
    TTL_FONT  = Font(name="Meiryo UI", color=NAVY,   bold=True, size=13)
    SEC_FONT  = Font(name="Meiryo UI", color=NAVY,   bold=True, size=10)

    THIN      = Side(style="thin")
    NO_SIDE   = Side(style=None)
    FULL_B    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    AL_CC  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_LC  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    AL_LT  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    # ── 列数・列幅 ──────────────────────────────────────
    NCOLS = 17   # A(1) 〜 Q(17)

    # 列幅: A=No/hdr, B-C=日付, D=業種, E=概要(広), F=OS/DB, G=役割, H-Q=フェーズ×10
    col_widths = {
        1: 6,    # A
        2: 9,    # B
        3: 9,    # C
        4: 10,   # D
        5: 42,   # E  ← 概要・担当業務（基本情報では値エリアに使用）
        6: 18,   # F
        7: 7,    # G
        8: 5,    # H
        9: 5,    # I
        10: 5,   # J
        11: 5,   # K
        12: 5,   # L
        13: 5,   # M
        14: 5,   # N
        15: 5,   # O
        16: 5,   # P
        17: 5,   # Q
    }
    from openpyxl.utils import get_column_letter
    for col_i, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col_i)].width = w

    # ── ページ設定（A4横）──────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6)

    # ════════════════════════════════════════════════════
    # ヘルパー：マージ前に外枠罫線を設定してからマージ
    # ════════════════════════════════════════════════════
    def _border_then_merge(r1, c1, r2, c2):
        """全セルに外枠罫線を設定 → マージ"""
        for row in range(r1, r2 + 1):
            for col in range(c1, c2 + 1):
                L = THIN if col == c1 else NO_SIDE
                R = THIN if col == c2 else NO_SIDE
                T = THIN if row == r1 else NO_SIDE
                B = THIN if row == r2 else NO_SIDE
                ws.cell(row=row, column=col).border = Border(
                    left=L, right=R, top=T, bottom=B)
        ws.merge_cells(
            start_row=r1, start_column=c1,
            end_row=r2,   end_column=c2)

    def _hdr(row, col, text):
        """単一ヘッダーセル（紺背景・白字）"""
        c = ws.cell(row=row, column=col, value=text)
        c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, AL_CC, FULL_B

    def _hdr_m(row, c1, c2, text):
        """マージヘッダーセル（紺背景・白字）"""
        _border_then_merge(row, c1, row, c2)
        c = ws.cell(row=row, column=c1, value=text)
        c.fill, c.font, c.alignment = HDR_FILL, HDR_FONT, AL_CC

    def _val_m(row, c1, c2, text, align=None, fill=None):
        """マージ値セル（白背景）"""
        _border_then_merge(row, c1, row, c2)
        c = ws.cell(row=row, column=c1, value=safe_str(text))
        c.font, c.alignment = VAL_FONT, (align or AL_LC)
        c.fill = fill or WHT_FILL

    def _val(row, col, text, align=None, fill=None):
        """単一値セル"""
        c = ws.cell(row=row, column=col, value=safe_str(text))
        c.font, c.alignment, c.border = VAL_FONT, (align or AL_LC), FULL_B
        c.fill = fill or WHT_FILL

    def _title_m(row, c1, c2, text, font, align=AL_CC, fill=None):
        """タイトル・セクション行（マージ）"""
        _border_then_merge(row, c1, row, c2)
        c = ws.cell(row=row, column=c1, value=text)
        c.font, c.alignment = font, align
        if fill: c.fill = fill

    def _auto_height(row, *texts, base=18, chars_per_line=38):
        max_lines = 1
        for t in texts:
            s = str(t or "")
            n = s.count("\n") + max(1, (len(s) + chars_per_line - 1) // chars_per_line)
            max_lines = max(max_lines, n)
        ws.row_dimensions[row].height = max(base, max_lines * 13)

    # ════════════════════════════════════════════════════
    # データ取得
    # ════════════════════════════════════════════════════
    bi = data.get("基本情報", {}) or {}
    si = data.get("SAP情報",  {}) or {}

    r = 1  # 現在行

    # ══ Row1: タイトル ══════════════════════════════════
    _title_m(r, 1, NCOLS, "職　務　経　歴　書", TTL_FONT)
    ws.row_dimensions[r].height = 30
    r += 1

    # ══ Row2〜4: 基本情報（全行で列区切りを統一） ════════
    #
    # 【列区切り統一】
    #   Col1      : ラベル（単一・紺）
    #   Col2-4    : 値①
    #   Col5      : ラベル（単一・紺）
    #   Col6-9    : 値②
    #   Col10     : ラベル（単一・紺）
    #   Col11-13  : 値③
    #   Col14     : ラベル（単一・紺）
    #   Col15-17  : 値④
    #
    # Row2: フリガナ | 値 | 性別 | 値 | 生年月日 | 値 | 未・既婚 | 値
    # Row3: 氏名    | 値 | 国籍  | 値 | 日本滞在  | 値 | SAP経験  | 値
    # Row4: モジュール | 値 | ポジション | 値 | 住所 | 値(span to end)

    # Row2
    _hdr(r, 1,  "フリガナ")
    _val_m(r, 2,  4,  bi.get("フリガナ"))
    _hdr(r, 5,  "性別")
    _val_m(r, 6,  9,  bi.get("性別"), AL_CC)
    _hdr(r, 10, "生年月日")
    _val_m(r, 11, 13, bi.get("生年月日"))
    _hdr(r, 14, "未・既婚")
    _val_m(r, 15, 17, bi.get("未既婚"), AL_CC)
    ws.row_dimensions[r].height = 18
    r += 1

    # Row3
    _hdr(r, 1,  "氏　名")
    _val_m(r, 2,  4,  bi.get("氏名"))
    _hdr(r, 5,  "国籍")
    _val_m(r, 6,  9,  bi.get("国籍"), AL_CC)
    _hdr(r, 10, "日本滞在")
    _val_m(r, 11, 13, bi.get("日本滞在年数"))
    _hdr(r, 14, "SAP経験")
    _val_m(r, 15, 17, si.get("SAP経験年数"), AL_CC)
    ws.row_dimensions[r].height = 18
    r += 1

    # Row4: モジュール / ポジション / 住所（住所は残り全列）
    _hdr(r, 1,  "モジュール")
    _val_m(r, 2,  4,  si.get("モジュール"))
    _hdr(r, 5,  "ポジション")
    _val_m(r, 6,  9,  si.get("ポジション"))
    _hdr(r, 10, "住　所")
    _val_m(r, 11, 17, bi.get("住所"))
    ws.row_dimensions[r].height = 18
    r += 1

    # Row5: 最寄駅
    _hdr(r, 1, "最寄駅")
    route   = safe_str(bi.get("最寄駅路線", ""), "")
    name    = safe_str(bi.get("最寄駅名",   ""), "")
    station = " ".join(filter(None, [route, (name + "駅") if name else ""])) \
              or safe_str(bi.get("最寄駅", ""))
    _val_m(r, 2, 17, station)
    ws.row_dimensions[r].height = 18
    r += 1

    # Row6: 取得資格
    _hdr(r, 1, "取得資格")
    _val_m(r, 2, 17, data.get("取得資格"), AL_LT)
    _auto_height(r, data.get("取得資格"), base=18, chars_per_line=90)
    r += 1

    # Row7: 得意分野
    _hdr(r, 1, "得意分野")
    _val_m(r, 2, 17, data.get("得意分野"), AL_LT)
    _auto_height(r, data.get("得意分野"), base=18, chars_per_line=90)
    r += 1

    # Row8: 自己PR
    _hdr(r, 1, "自己ＰＲ")
    _val_m(r, 2, 17, data.get("自己PR"), AL_LT)
    _auto_height(r, data.get("自己PR"), base=36, chars_per_line=50)
    r += 1

    # 区切り空行（罫線なし・白）
    ws.row_dimensions[r].height = 6
    r += 1

    # ══ 業務経歴タイトル ════════════════════════════════
    _title_m(r, 1, NCOLS, "業　務　経　歴", SEC_FONT)
    ws.row_dimensions[r].height = 22
    r += 1

    # ══ 業務経歴ヘッダー ════════════════════════════════
    proj_header_row = r
    for ci, hv in enumerate([
        "No", "開始\n年月", "終了\n年月", "業種",
        "システム概要・担当業務", "OS/言語/DB/ツール", "役割",
        "分析\n調査", "提案\n管理", "要件\n定義", "基本\n設計",
        "詳細\n設計", "製造", "単体\n試験", "結合\n試験", "総合\n試験", "運用\n保守"
    ], 1):
        _hdr(r, ci, hv)
    ws.row_dimensions[r].height = 30
    r += 1

    # ══ 業務経歴データ ══════════════════════════════════
    phase_keys = [
        "分析調査", "提案管理レビュー", "要件定義", "基本設計",
        "詳細設計", "製造", "単体試験", "結合試験", "総合試験", "運用保守"
    ]

    for pi, proj in enumerate(data.get("職務経歴", []) or []):
        phases  = proj.get("フェーズ", {}) or {}
        row_bg  = STR_FILL if pi % 2 == 0 else WHT_FILL

        overview = safe_str(proj.get("プロジェクト概要"))
        tasks    = safe_str(proj.get("担当業務"))
        combined = (f"【概要】{overview}\n【担当】{tasks}"
                    if tasks not in ("-", overview, "") else overview)

        os_db    = safe_str(proj.get("OS_DB"))
        env      = safe_str(proj.get("作業環境"))
        lang     = safe_str(proj.get("開発言語"))
        env_text = "\n".join(x for x in [os_db, env, lang] if x and x != "-") or "-"

        row_data = [
            safe_str(proj.get("No", pi + 1)),
            safe_str(proj.get("開始年月")),
            safe_str(proj.get("終了年月")),
            safe_str(proj.get("業種")),
            combined,
            env_text,
            safe_str(proj.get("役割")),
        ] + ["●" if phases.get(k) else "" for k in phase_keys]

        for ci, v in enumerate(row_data, 1):
            al = AL_CC if (ci <= 3 or ci >= 8) else (AL_LT if ci in (5, 6) else AL_LC)
            _val(r, ci, v, al, row_bg)

        _auto_height(r, combined, env_text, base=25, chars_per_line=36)
        r += 1

    # ── フリーズ・印刷範囲 ────────────────────────────
    ws.freeze_panes = ws.cell(row=proj_header_row + 1, column=1)
    ws.print_area   = f"A1:Q{r - 1}"

    buf = _BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────
# メインUI
# ─────────────────────────────────────────

st.markdown('<div class="main-title">📄 スキルシート変換ツール</div>', unsafe_allow_html=True)
st.caption("各社バラバラのExcel/PDFスキルシートを自社フォーマットに自動変換します")

st.divider()

# サイドバー：API設定
with st.sidebar:
    st.header("⚙️ 設定")
    api_key = st.text_input(
        "Gemini APIキー",
        value=GEMINI_API_KEY,
        type="password",
        help="Google AI StudioのAPIキーを入力"
    )
    st.caption("🔒 キーはこのセッション内のみ使用されます")
    st.divider()
    st.markdown("**対応ファイル形式**")
    st.markdown("- 📊 Excel (.xlsx, .xls)\n- 📄 PDF (.pdf)")
    st.divider()
    st.markdown("**処理フロー**")
    st.markdown("1. ファイルアップロード\n2. テキスト抽出\n3. Gemini AI解析\n4. 自社フォーマット表示\n5. PDF出力")

# メインエリア
col1, col2 = st.columns([1, 1])

with col1:
    st.markdown('<div class="section-header">📁 ファイルアップロード</div>', unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        "スキルシートをアップロード（複数可）",
        type=["xlsx", "xls", "pdf"],
        accept_multiple_files=True,
        help="ExcelまたはPDF形式のスキルシートをドラッグ＆ドロップ"
    )

    if uploaded_files:
        st.success(f"✅ {len(uploaded_files)}件のファイルが選択されました")
        for f in uploaded_files:
            size_kb = len(f.getvalue()) / 1024
            st.caption(f"• {f.name}（{size_kb:.1f} KB）")

with col2:
    st.markdown('<div class="section-header">🚀 変換実行</div>', unsafe_allow_html=True)
    if not api_key:
        st.warning("⚠️ サイドバーにGemini APIキーを入力してください")
    elif not uploaded_files:
        st.info("👆 左側からファイルをアップロードしてください")
    else:
        if st.button("🤖 AIで自動変換する", type="primary", use_container_width=True):
            st.session_state["results"] = []
            for uploaded_file in uploaded_files:
                with st.spinner(f"⏳ {uploaded_file.name} を処理中..."):
                    try:
                        file_bytes = uploaded_file.getvalue()
                        ext = uploaded_file.name.lower().split(".")[-1]

                        # テキスト抽出
                        if ext in ["xlsx", "xls"]:
                            text = extract_text_from_excel(file_bytes, ext)
                        else:
                            text = extract_text_from_pdf(file_bytes)

                        # Gemini API呼び出し
                        extracted = extract_skills_with_gemini(text, api_key)
                        extracted["_filename"] = uploaded_file.name
                        st.session_state["results"].append(extracted)
                        st.success(f"✅ {uploaded_file.name} 完了！")
                    except json.JSONDecodeError as e:
                        st.error(f"❌ JSON解析エラー: {uploaded_file.name}\n{e}")
                    except Exception as e:
                        st.error(f"❌ エラー: {uploaded_file.name}\n{e}")

# ─────────────────────────────────────────
# 結果表示エリア
# ─────────────────────────────────────────
if "results" in st.session_state and st.session_state["results"]:
    st.divider()
    st.markdown("## 📊 抽出結果・プレビュー")

    for idx, data in enumerate(st.session_state["results"]):
        bi = data.get("基本情報", {})
        si = data.get("SAP情報", {})
        name = bi.get("氏名", f"ファイル{idx+1}")

        with st.expander(f"👤 {name}（{data.get('_filename','')}）", expanded=True):

            # 基本情報
            st.markdown('<div class="section-header">👤 基本情報</div>', unsafe_allow_html=True)
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("氏名", bi.get("氏名", "-"))
            c2.metric("生年月日", bi.get("生年月日", "-"))
            c3.metric("性別", bi.get("性別", "-"))
            c4.metric("未・既婚", bi.get("未既婚", "-"))

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("モジュール", si.get("モジュール", "-"))
            c2.metric("ポジション", si.get("ポジション", "-"))
            c3.metric("SAP経験年数", si.get("SAP経験年数", "-"))
            c4.metric("最寄駅", f"{bi.get('最寄駅路線','')} {bi.get('最寄駅名','')}駅")

            # 資格・PR
            st.markdown('<div class="section-header">📜 資格・スキル</div>', unsafe_allow_html=True)
            st.info(f"**取得資格：** {data.get('取得資格', '-')}")

            with st.container():
                st.markdown("**得意分野**")
                st.text_area("得意分野", data.get("得意分野", ""), height=80,
                              label_visibility="collapsed", key=f"tokui_{idx}")
                st.markdown("**自己PR**")
                st.text_area("自己PR", data.get("自己PR", ""), height=80,
                              label_visibility="collapsed", key=f"pr_{idx}")

            # 職務経歴
            st.markdown('<div class="section-header">💼 職務経歴</div>', unsafe_allow_html=True)
            projects = data.get("職務経歴", [])
            if projects:
                for proj in projects:
                    phases = proj.get("フェーズ", {})
                    active = [k for k, v in phases.items() if v]
                    st.markdown(f"""
<div class="project-card">
<b>No.{proj.get('No','')} | {proj.get('開始年月','')} ～ {proj.get('終了年月','')}</b>
　<span class="badge">{proj.get('業種','')}</span>
　<span class="badge">{proj.get('役割','')}</span>
<br><br>
{proj.get('プロジェクト概要','')}
<br><br>
<b>担当：</b>{proj.get('担当業務','')}
<br>
<b>環境：</b>{proj.get('OS_DB','')} / {proj.get('作業環境','')} / {proj.get('開発言語','')}
<br>
<b>フェーズ：</b>{"　".join([f'<span class="skill-tag">{p}</span>' for p in active])}
</div>
""", unsafe_allow_html=True)
            else:
                st.warning("職務経歴が抽出できませんでした")

            # Excelダウンロードボタン
            st.markdown("---")
            try:
                excel_bytes = generate_excel(data)
                safe_name = safe_str(bi.get("氏名", f"skillsheet_{idx}"), f"skillsheet_{idx}").replace(" ", "_")
                st.download_button(
                    label="📊 Excelダウンロード（自社フォーマット）",
                    data=excel_bytes,
                    file_name=f"{safe_name}_職務経歴書.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                    key=f"xl_{idx}"
                )
            except Exception as e:
                st.error(f"Excel生成エラー: {e}")
                import traceback
                st.code(traceback.format_exc(), language="text")

    # 一括PDF出力
    if len(st.session_state["results"]) > 1:
        st.divider()
        st.markdown("### 📦 一括ダウンロード")
        st.info("複数ファイルは個別にExcelダウンロードしてください（上記ボタン）")
